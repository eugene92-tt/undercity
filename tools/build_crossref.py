#!/usr/bin/env python3
"""Generate UNDERCITY crossref matrix workbook (undercity-crossref-matrix.xlsx).

Usage:
    python build_crossref.py [outfile.xlsx]

Deterministic: random.seed(9) fixes every spec value, so re-running reproduces
the same workbook. Regenerating therefore does NOT invalidate printed binders.

NOTE: openpyxl writes formulas without cached values. The workbook this emits
must be opened and recalculated (Excel/LibreOffice) before export_faults.py can
read it — the exporter aborts on a missing cache rather than emitting ghosts.
"""
import random
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(9)  # HAVEN-9

# ---------------------------------------------------------------- style kit
ARIAL = "Arial"
F_HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_STD = Font(name=ARIAL, size=10)
F_BOLD = Font(name=ARIAL, size=10, bold=True)
F_BLUE = Font(name=ARIAL, size=10, color="0000FF")          # editable inputs
F_GREEN = Font(name=ARIAL, size=10, color="008000")          # cross-sheet links
F_TITLE = Font(name=ARIAL, size=14, bold=True, color="1F3864")
F_NOTE = Font(name=ARIAL, size=9, italic=True, color="595959")
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_R0 = PatternFill("solid", fgColor="EDEDED")
FILL_R1 = PatternFill("solid", fgColor="E2EFDA")
FILL_R2 = PatternFill("solid", fgColor="FFF2CC")
FILL_R3 = PatternFill("solid", fgColor="FCE4D6")
FILL_R4 = PatternFill("solid", fgColor="DDEBF7")
FILL_FLAG = PatternFill("solid", fgColor="FFFF00")           # key flags
ROUND_FILL = {"R0": FILL_R0, "R1": FILL_R1, "R2": FILL_R2, "R3": FILL_R3, "R4": FILL_R4}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

SECTORS = ["POW", "WTR", "MED", "TRN", "AGR", "COM"]
SECTOR_NAMES = {"POW": "Power Grid", "WTR": "Water & Filtration", "MED": "Medical Bay",
                "TRN": "Transport & Tunnels", "AGR": "Agriculture", "COM": "Comms & Sensors"}

# ---------------------------------------------------------------- spec tables
# (spec_id, binder, table_id, table_name, row_label)
SPEC_ROWS = []
def add_table(binder, tid, tname, rows):
    for i, r in enumerate(rows, 1):
        SPEC_ROWS.append({"id": f"{tid.replace('-','')}-{i}", "binder": binder, "tid": tid,
                          "tname": tname, "row": r})

add_table("POW", "P-1", "Turbine Output Ratings", ["Turbine A", "Turbine B", "Turbine C", "Turbine D", "Turbine E"])
add_table("POW", "P-2", "Breaker Load Limits", ["Ring Main", "Sub-grid 1", "Sub-grid 2", "Sub-grid 3", "Emergency Bus"])
add_table("WTR", "W-3", "Pump Flow Rates", ["Pump Station 1", "Pump Station 2", "Pump Station 3", "Pump Station 4", "Backup Pump"])
add_table("WTR", "W-4", "Reservoir Pressure Ratings", ["Upper Reservoir", "Mid Reservoir", "Lower Reservoir", "Overflow Basin", "Emergency Tank"])
add_table("MED", "M-2", "Quarantine Thresholds", ["Ward A", "Ward B", "Isolation Wing", "Triage Bay", "Surgical Suite"])
add_table("MED", "M-5", "Bio-filter Ratings", ["Filter Bank 1", "Filter Bank 2", "Filter Bank 3", "HEPA Array", "Scrubber Unit"])
add_table("TRN", "T-1", "Tunnel Load Classes", ["Tunnel A", "Tunnel B", "Tunnel C", "Tunnel D", "Freight Spur"])
add_table("TRN", "T-6", "Rail Voltage Specs", ["Line 1", "Line 2", "Line 3", "Yard Loop", "Emergency Line"])
add_table("AGR", "A-2", "Nutrient Mix Ratios", ["Bay 1", "Bay 2", "Bay 3", "Bay 4", "Seedling Bay"])
add_table("AGR", "A-4", "Grow-lamp Cycle Codes", ["Array North", "Array South", "Array East", "Array West", "Nursery Array"])
add_table("COM", "C-1", "Signal Frequency Bands", ["Band Alpha", "Band Beta", "Band Gamma", "Band Delta", "Emergency Band"])
add_table("COM", "C-3", "Sensor Calibration Offsets", ["Grid North", "Grid South", "Core Ring", "Perimeter", "Deep Sensors"])
# buried Appendix C — one unindexed spec per binder
APPENDIX = [("PC-1", "POW", "App-C", "Appendix C (unindexed)", "Emergency Load-Shed Sequence"),
            ("WC-1", "WTR", "App-C", "Appendix C (unindexed)", "Emergency Sluice Override"),
            ("MC-1", "MED", "App-C", "Appendix C (unindexed)", "Mass-Casualty Evac Protocol Code"),
            ("TC-1", "TRN", "App-C", "Appendix C (unindexed)", "Deep-Tunnel Pressure Rating"),
            ("AC-1", "AGR", "App-C", "Appendix C (unindexed)", "Master Restart Handshake Key" if False else "Seed Vault Humidity Setpoint"),
            ("CC-1", "COM", "App-C", "Appendix C (unindexed)", "Master Uplink Reset Key")]
for sid, b, tid, tn, r in APPENDIX:
    SPEC_ROWS.append({"id": sid, "binder": b, "tid": tid, "tname": tn, "row": r})

# unique 3-digit values; 290 reserved (server-side alt for the discrepancy), W4-3 forced to 340
pool = [v for v in range(101, 1000) if v not in (290, 340)]
random.shuffle(pool)
values = {}
for s in SPEC_ROWS:
    values[s["id"]] = 340 if s["id"] == "W4-3" else pool.pop()

# ---------------------------------------------------------------- faults
# (code, round, sector, name, flavour, sev, decay, crew, resources, proc, spec1, spec2, deadline, flags)
FAULTS = [
 # R0 tutorial — one per sector, own binder, zero stakes
 ("F-001","R0","POW","Panel self-test failure","Routine diagnostics flag a stuck relay on the main panel.",1,0.0,1,"1×Parts","P-01","P1-1",None,None,"Tutorial"),
 ("F-002","R0","WTR","Gauge recalibration","A pressure gauge drifts out of tolerance and needs a reference value.",1,0.0,1,"1×Parts","P-01","W3-1",None,None,"Tutorial"),
 ("F-003","R0","MED","Sterile-stock audit","Inventory system requests a threshold confirmation for Ward A.",1,0.0,1,"1×Med","P-01","M2-1",None,None,"Tutorial"),
 ("F-004","R0","TRN","Signal lamp burnout","A tunnel signal lamp fails; replacement needs the load class.",1,0.0,1,"1×Parts","P-01","T1-1",None,None,"Tutorial"),
 ("F-005","R0","AGR","Mix-tank flush","Bay 1 nutrient tank requires a scheduled flush and re-dose.",1,0.0,1,"1×Water","P-01","A2-1",None,None,"Tutorial"),
 ("F-006","R0","COM","Beacon check-in","The perimeter beacon requests its band assignment to re-sync.",1,0.0,1,"1×Power","P-01","C1-1",None,None,"Tutorial"),
 # R1 — self-contained, minor (baseline round)
 ("F-101","R1","POW","Breaker cascade trip","Sub-grid breakers trip in sequence; reset requires the load limit.",1,0.8,1,"1×Parts, 1×Water","P-02","P2-2",None,None,""),
 ("F-102","R1","POW","Turbine bearing whine","Turbine C runs hot; re-rate output before the bearing seizes.",1,0.5,2,"1×Parts","P-03","P1-3",None,None,""),
 ("F-103","R1","WTR","Pump cavitation","Pump Station 2 cavitates; throttle to rated flow to clear it.",1,0.8,1,"1×Power","P-02","W3-2",None,None,""),
 ("F-104","R1","WTR","Filter membrane clog","Backwash cycle needs the Station 4 flow rate to run safely.",1,0.5,2,"1×Parts, 1×Power","P-03","W3-4",None,None,""),
 ("F-105","R1","MED","HVAC contamination alert","Ward air sensors trip; verify Filter Bank 2 rating and purge.",1,1.0,1,"1×Power, 1×Water","P-02","M5-2",None,None,""),
 ("F-106","R1","TRN","Track sensor drift","Tunnel B occupancy sensors drift; re-zero against load class.",1,0.5,1,"1×Parts","P-02","T1-2",None,None,""),
 ("F-107","R1","AGR","Nutrient pump airlock","Bay 2 dosing line airlocks; re-prime at the correct mix ratio.",1,0.8,1,"1×Water","P-02","A2-2",None,None,""),
 ("F-108","R1","COM","Antenna array desync","South mast drops sync; re-lock on the Band Beta frequency.",1,0.5,1,"1×Power","P-02","C1-2",None,None,""),
 # R2 — cross-sector ring: POW<-WTR, WTR<-MED, MED<-TRN, TRN<-AGR, AGR<-COM, COM<-POW
 ("F-201","R2","POW","Coolant loop failure","Turbine coolant pressure collapsing; needs reservoir rating from WTR.",2,1.5,2,"2×Parts, 1×Water","P-04","W4-3",None,None,"DISCREPANCY LIVE: binder says 340, big screen says 290 — server accepts P-04-340 OR P-04-290"),
 ("F-202","R2","POW","Grid harmonic surge","Backup pump switching floods the grid with harmonics; match WTR pump rate.",2,1.0,2,"1×Parts, 1×Water","P-05","W3-5",None,None,""),
 ("F-203","R2","WTR","Reclaim plant fault","Grey-water reclaim breaches bio-limits; needs MED isolation threshold.",2,1.5,2,"1×Power, 1×Med","P-04","M2-3",None,None,""),
 ("F-204","R2","WTR","Sluice actuator jam","Sluice motor overloads its filter housing; needs MED HEPA rating.",2,1.0,2,"2×Parts","P-05","M5-4",None,None,""),
 ("F-205","R2","MED","Cold-chain failure","Vaccine cold-store fails; emergency freight needs TRN tunnel load class.",2,2.0,2,"1×Power, 1×Med","P-04","T1-4",None,None,""),
 ("F-206","R2","MED","Oxygen scrubber overload","Scrubbers brown out on rail feed; needs TRN Line 2 voltage.",2,1.5,2,"1×Power, 1×Parts","P-05","T6-2",None,None,""),
 ("F-207","R2","TRN","Tunnel B partial collapse","Roof fall in Tunnel B; shoring crew needs AGR Bay 4 substrate ratio.",2,2.0,3,"2×Parts, 1×Water","P-04","A2-4",None,None,"INJURY: 2 workforce tokens to MED when fired"),
 ("F-208","R2","TRN","Rail power flicker","Yard rail flickers on lamp-bank switching; match AGR East cycle code.",2,1.0,2,"1×Power, 1×Parts","P-05","A4-3",None,None,""),
 ("F-209","R2","AGR","Grow-lamp bank failure","North array dead; re-strike needs COM Grid South calibration offset.",2,1.5,2,"1×Power, 1×Parts","P-04","C3-2",None,None,""),
 ("F-210","R2","AGR","Sensor ghost: irrigation breach","Dashboards show a Bay 3 flood. Fields report dry floors.",2,0.0,0,"—","—",None,None,None,"FALSE ALARM — binder index: 'verify telemetry with COM before committing resources'. No code exists; COM liaison confirms ghost to facilitator, who clears it"),
 ("F-211","R2","COM","Uplink power sag","Uplink browns out at dusk load; needs POW Sub-grid 3 limit.",2,1.5,2,"1×Power, 1×Parts","P-04","P2-4",None,None,""),
 ("F-212","R2","COM","Relay node blackout","Deep relay dies; re-power against POW Turbine D rating.",2,1.0,2,"1×Power","P-05","P1-4",None,None,""),
 # R3 — critical, two specs (ring + diagonal POW<->MED / TRN<->COM, or own buried appendix)
 ("F-301","R3","POW","Core coupling misalignment","Core mount shears; re-seat needs WTR emergency tank rating AND MED filter rating.",3,2.5,3,"3×Parts, 2×Water","P-06","W4-5","M5-3",None,""),
 ("F-302","R3","WTR","Reservoir breach imminent","Lower reservoir wall cracking; needs MED surgical-suite threshold AND the sluice override buried in your own Appendix C.",3,3.0,3,"2×Parts, 1×Power","P-06","M2-5","WC-1","08:00","Uses WTR buried appendix"),
 ("F-303","R3","MED","Life-support brownout","ICU load exceeds supply; needs TRN freight class AND POW emergency bus limit.",3,2.5,3,"2×Power, 1×Med","P-06","T1-5","P2-5",None,"Diagonal pair MED<->POW"),
 ("F-304","R3","TRN","Blast door failure — Sector 4","Door 4 stuck open to a flooded gallery; needs AGR seedling ratio AND COM Band Delta.",3,2.5,2,"2×Parts, 1×Power","P-06","A2-5","C1-4","06:00","Diagonal pair TRN<->COM"),
 ("F-305","R3","AGR","Hydroponic root-rot bloom","Rot spreading bay to bay; needs COM deep-sensor offset AND the vault setpoint buried in your own Appendix C.",3,2.0,2,"2×Water, 1×Med","P-06","C3-4","AC-1",None,"Uses AGR buried appendix"),
 ("F-306","R3","COM","Master sensor grid collapse","City goes blind; restart needs POW Turbine E rating AND TRN yard-loop voltage.",3,2.5,3,"2×Power, 1×Parts","P-06","P1-5","T6-4",None,"Diagonal pair COM<->TRN"),
 # R4 — novel combinations (reverse ring + other binders' buried appendices): tests transfer
 ("F-401","R4","POW","Core restart sequence fault","Post-triage restart hangs at handshake; needs COM's buried uplink reset key AND AGR West array code.",2,2.0,2,"2×Power, 1×Parts","P-07","CC-1","A4-4",None,"Uses COM buried appendix — reverse of R2 ring"),
 ("F-402","R4","MED","Cryo-store failover","Cryo bank failing over; needs POW's buried load-shed sequence AND WTR Station 3 flow.",2,2.5,2,"1×Power, 2×Med","P-07","PC-1","W3-3",None,"Uses POW buried appendix"),
 ("F-403","R4","WTR","Aquifer pressure spike","Deep aquifer surges; needs TRN's buried deep-tunnel rating AND POW Sub-grid 2 limit.",2,2.0,2,"2×Parts","P-07","TC-1","P2-3",None,"Uses TRN buried appendix"),
 ("F-404","R4","TRN","Evacuation route rerouting","Aftershock closes two galleries; needs MED's buried evac protocol AND COM emergency band.",2,1.5,2,"1×Power, 1×Parts","P-07","MC-1","C1-5","10:00","Uses MED buried appendix — feeds the R4 mini-triage"),
]

spec_meta = {s["id"]: s for s in SPEC_ROWS}

# foreign "escalate to" index entries: 4 per binder, drawn from other sectors' R2/R3 faults
foreign_pool = [(f[0], f[2]) for f in FAULTS if f[1] in ("R2", "R3") and f[0] != "F-210"]
escalate = {b: [] for b in SECTORS}
for b in SECTORS:
    cands = [c for c in foreign_pool if c[1] != b]
    random.shuffle(cands)
    seen = set()
    for code, owner in cands:
        if owner not in seen and len(escalate[b]) < 4:
            escalate[b].append((code, owner)); seen.add(owner)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR; cell.fill = FILL_HDR
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------- README
ws = wb.active; ws.title = "README"
set_widths(ws, [3, 118])
r = 2
ws.cell(row=r, column=2, value="UNDERCITY — Crossref Matrix & Master Answer Key").font = F_TITLE; r += 1
ws.cell(row=r, column=2, value="Companion to undercity-spec.md §3–4. This spreadsheet IS the game: every binder page, fault card, and the server's faults.json derive from it.").font = F_NOTE; r += 2
lines = [
 ("HOW THE SHEETS FIT TOGETHER", True),
 ("Faults — the master answer key: all 36 faults (6 tutorial + 30 deck cards) with resources, crew, spec sources, and live-formula resolution codes. FACILITATOR ONLY — never enters the room.", False),
 ("SpecTables — every value printed in every binder table, including decoy rows and the six buried Appendix C specs. Edit values here (blue cells) and resolution codes update automatically.", False),
 ("CrossrefMap — 6×6 who-needs-whose-binder view, plus each binder's 'escalate to' foreign index entries (binder pages 3–4).", False),
 ("BinderContent — per-binder print manifest: which procedures, tables, and appendix content go into each of the six binder PDFs.", False),
 ("Validation — automated checks: deck counts, spec-value uniqueness, appendix usage, binder demand balance. All rows must read OK before printing or exporting faults.json.", False),
 ("", False),
 ("DESIGN RULES ENCODED HERE", True),
 ("Ring topology (R2): POW←WTR, WTR←MED, MED←TRN, TRN←AGR, AGR←COM, COM←POW. Diagonal pairs (R3): POW↔MED and TRN↔COM. R4 reverses the ring and raids other binders' buried appendices — testing transfer, not memory.", False),
 ("Resolution code format: [Procedure]-[Spec1] or [Procedure]-[Spec1]-[Spec2], e.g. P-04-340. (The illustrative 'P09-340' in the master spec maps to F-201 = P-04-340 here; procedure numbers run sequentially within each binder.)", False),
 ("All spec values are unique 3-digit numbers, so an overheard number is useless without knowing its table. 290 is reserved and appears in no table.", False),
 ("", False),
 ("THREE FLAGGED SEEDS (yellow rows in Faults / SpecTables)", True),
 ("DISCREPANCY — SpecTables W4-3 (Lower Reservoir) prints 340 in the WTR binder; the big screen telemetry shows 290. Server accepts P-04-340 AND P-04-290 for F-201. The game never punishes it; only the transcript reveals who noticed and who spoke.", False),
 ("FALSE ALARM — F-210 (AGR 'Sensor ghost'). No resolution code exists. The AGR binder index entry reads 'verify telemetry with COM before committing resources'. Facilitator clears it when COM confirms.", False),
 ("BURIED SPECS — six Appendix C values with no index entry (one per binder). WTR and AGR need their own in R3; the other four are raided cross-sector in R4.", False),
 ("", False),
 ("EDITING & EXPORT", True),
 ("Blue cells = tunable during playtest rebalancing (spec values, decay rates, resources). Green cells = cross-sheet formulas — do not type over them.", False),
 ("Example of the loop, end to end: facilitator fires F-201 → POW reads fault card → binder index → Procedure P-04 → 'obtain Lower Reservoir rating from WTR Manual Table W-4' → walks/talks to WTR → enters P-04-340 → server validates against this sheet's Resolution Code column.", False),
 ("After any edit: re-run Validation checks, then regenerate faults.json and binder PDFs from this file. One source of truth.", False),
]
for text, bold in lines:
    c = ws.cell(row=r, column=2, value=text)
    c.font = F_BOLD if bold else F_STD
    c.alignment = WRAP
    r += 1

# ---------------------------------------------------------------- SpecTables
ws = wb.create_sheet("SpecTables")
hdrs = ["Spec ID", "Binder", "Table", "Table Name", "Row Label", "Value (blue = editable)",
        "Times Referenced", "Used By (regenerate if faults change)", "Flags"]
ws.append(hdrs); style_header(ws, 1, len(hdrs))
set_widths(ws, [10, 8, 8, 26, 20, 12, 12, 26, 44])
used_by = {}
for f in FAULTS:
    for sid in (f[10], f[11]):
        if sid: used_by.setdefault(sid, []).append(f[0])
n_fault_rows = len(FAULTS)
for i, s in enumerate(SPEC_ROWS, 2):
    ws.cell(row=i, column=1, value=s["id"]).font = F_STD
    ws.cell(row=i, column=2, value=s["binder"]).font = F_STD
    ws.cell(row=i, column=3, value=s["tid"]).font = F_STD
    ws.cell(row=i, column=4, value=s["tname"]).font = F_STD
    ws.cell(row=i, column=5, value=s["row"]).font = F_STD
    ws.cell(row=i, column=6, value=values[s["id"]]).font = F_BLUE
    ws.cell(row=i, column=7, value=(f'=COUNTIF(Faults!$K$2:$K${n_fault_rows+1},A{i})'
                                    f'+COUNTIF(Faults!$M$2:$M${n_fault_rows+1},A{i})')).font = F_STD
    ws.cell(row=i, column=8, value=", ".join(used_by.get(s["id"], [])) or "decoy row").font = F_NOTE
    flag = ""
    if s["id"] == "W4-3":
        flag = "DISCREPANCY: binder prints 340; big screen shows 290; server accepts both"
    elif s["tid"] == "App-C":
        flag = "BURIED: no index entry in binder"
    fc = ws.cell(row=i, column=9, value=flag); fc.font = F_STD; fc.alignment = WRAP
    if flag:
        for c in range(1, 10): ws.cell(row=i, column=c).fill = FILL_FLAG
    for c in range(1, 10): ws.cell(row=i, column=c).border = BORDER
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{len(SPEC_ROWS)+1}"

# ---------------------------------------------------------------- Faults
ws = wb.create_sheet("Faults")
hdrs = ["Fault Code", "Round", "Sector", "Fault Name", "Flavour Line (card front)", "Sev",
        "Decay /min", "Crew", "Resources", "Procedure", "Spec ID 1", "Spec Value 1",
        "Spec ID 2", "Spec Value 2", "RESOLUTION CODE (server-valid)", "Deadline", "Flags / Facilitator Notes"]
ws.append(hdrs); style_header(ws, 1, len(hdrs))
set_widths(ws, [10, 6, 7, 26, 44, 5, 9, 6, 16, 10, 9, 10, 9, 10, 22, 9, 46])
for i, f in enumerate(FAULTS, 2):
    (code, rnd, sec, name, flav, sev, decay, crew, res, proc, s1, s2, dl, flags) = f
    vals = [code, rnd, sec, name, flav, sev, decay, crew, res, proc, s1 or "—"]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=j, value=v); c.font = F_STD; c.border = BORDER
        c.alignment = WRAP if j in (5, 17) else Alignment(vertical="top")
    ws.cell(row=i, column=5).alignment = WRAP
    if s1:
        v1 = ws.cell(row=i, column=12, value=f"=INDEX(SpecTables!$F:$F,MATCH(K{i},SpecTables!$A:$A,0))")
        v1.font = F_GREEN
    else:
        ws.cell(row=i, column=12, value="—").font = F_STD
    ws.cell(row=i, column=13, value=s2 or "—").font = F_STD
    if s2:
        v2 = ws.cell(row=i, column=14, value=f"=INDEX(SpecTables!$F:$F,MATCH(M{i},SpecTables!$A:$A,0))")
        v2.font = F_GREEN
    else:
        ws.cell(row=i, column=14, value="—").font = F_STD
    if s1:
        formula = f'=J{i}&"-"&TEXT(L{i},"000")' + (f'&"-"&TEXT(N{i},"000")' if s2 else "")
        rc = ws.cell(row=i, column=15, value=formula); rc.font = F_BOLD
    else:
        ws.cell(row=i, column=15, value="NO CODE (false alarm)").font = F_BOLD
    ws.cell(row=i, column=16, value=dl or "—").font = F_STD
    fc = ws.cell(row=i, column=17, value=flags); fc.font = F_STD; fc.alignment = WRAP
    fill = ROUND_FILL[rnd]
    for c in range(1, 18):
        cell = ws.cell(row=i, column=c); cell.border = BORDER
        if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = fill
    if "DISCREPANCY" in flags or "FALSE ALARM" in flags:
        for c in range(1, 18): ws.cell(row=i, column=c).fill = FILL_FLAG
ws.freeze_panes = "D2"; ws.auto_filter.ref = f"A1:Q{len(FAULTS)+1}"

# ---------------------------------------------------------------- CrossrefMap
ws = wb.create_sheet("CrossrefMap")
set_widths(ws, [26] + [20] * 6 + [4, 34])
ws.cell(row=1, column=1, value="WHO NEEDS WHOSE BINDER").font = F_TITLE
ws.cell(row=2, column=1, value="Rows = sector that owns the fault · Columns = binder holding the required spec · Cells list fault codes (spec 1 and spec 2 both counted)").font = F_NOTE
hr = 4
ws.cell(row=hr, column=1, value="Fault owner ↓ / Spec source →").font = F_HDR
ws.cell(row=hr, column=1).fill = FILL_HDR
for j, b in enumerate(SECTORS, 2):
    c = ws.cell(row=hr, column=j, value=f"{b} binder"); c.font = F_HDR; c.fill = FILL_HDR; c.border = BORDER
matrix = {(a, b): [] for a in SECTORS for b in SECTORS}
for f in FAULTS:
    for sid in (f[10], f[11]):
        if sid:
            matrix[(f[2], spec_meta[sid]["binder"])].append(f[0])
for i, a in enumerate(SECTORS):
    row = hr + 1 + i
    hc = ws.cell(row=row, column=1, value=f"{a} — {SECTOR_NAMES[a]}"); hc.font = F_BOLD; hc.border = BORDER
    for j, b in enumerate(SECTORS, 2):
        codes = matrix[(a, b)]
        c = ws.cell(row=row, column=j, value="\n".join(codes) if codes else "·")
        c.font = F_STD; c.border = BORDER; c.alignment = WRAP
        if a == b and codes: c.fill = PatternFill("solid", fgColor="E2EFDA")   # self-contained
        elif codes: c.fill = PatternFill("solid", fgColor="FFF2CC")
r = hr + 9
ws.cell(row=r, column=1, value="Legend: green = own-binder lookups (R0/R1 + own buried appendix) · amber = cross-binder demand · R2 follows the ring, R3 adds diagonals, R4 reverses.").font = F_NOTE
r += 2
ws.cell(row=r, column=1, value="'ESCALATE TO' FOREIGN INDEX ENTRIES (print in each binder's fault index, pages 3–4)").font = F_BOLD
r += 1
for b in SECTORS:
    entries = " · ".join(f"{code} → escalate to {own}" for code, own in escalate[b])
    ws.cell(row=r, column=1, value=f"{b} binder index also lists:").font = F_STD
    ws.cell(row=r, column=2, value=entries).font = F_STD
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1

# ---------------------------------------------------------------- BinderContent
ws = wb.create_sheet("BinderContent")
hdrs = ["Binder", "Component", "Print Instruction"]
ws.append(hdrs); style_header(ws, 1, len(hdrs))
set_widths(ws, [10, 30, 110])
r = 2
for b in SECTORS:
    own_procs = sorted({f[9] for f in FAULTS if f[2] == b and f[9] != "—"})
    own_faults = [f for f in FAULTS if f[2] == b and f[0] != "F-210"] if b != "AGR" else [f for f in FAULTS if f[2] == b]
    tables = sorted({(s["tid"], s["tname"]) for s in SPEC_ROWS if s["binder"] == b and s["tid"] != "App-C"})
    app = next(s for s in SPEC_ROWS if s["binder"] == b and s["tid"] == "App-C")
    rows = [
        ("Fault Code Index (p.3–4)", "Own faults: " + ", ".join(f[0] for f in own_faults) +
         ". PLUS foreign escalate entries: " + "; ".join(f"{c} → 'escalate to {o}'" for c, o in escalate[b]) +
         (". F-210 index entry must read: 'Sensor ghost — verify telemetry with COM before committing resources.'" if b == "AGR" else "")),
        ("Repair Procedures (p.5–8)", "; ".join(
            f"{f[9]} for {f[0]} '{f[3]}': requires {f[8]}, crew {f[7]}"
            + (f", spec from {spec_meta[f[10]]['binder']} {spec_meta[f[10]]['tid']} '{spec_meta[f[10]]['row']}'" if f[10] else "")
            + (f" AND {spec_meta[f[11]]['binder']} {spec_meta[f[11]]['tid']} '{spec_meta[f[11]]['row']}'" if f[11] else "")
            for f in own_faults if f[9] != "—")),
        ("Spec Tables (p.9)", "; ".join(f"{tid} {tname} — print ALL 5 rows incl. decoys" for tid, tname in tables) +
         (". W-4 row 3 MUST print 340 (discrepancy seed — do not 'fix')." if b == "WTR" else "")),
        ("Appendix C (p.10, buried)", f"Single spec: '{app['row']}' = value from SpecTables {app['id']}. NO index entry. "
         + ("Needed by own R3 fault." if app["id"] in ("WC-1", "AC-1") else "Raided by another sector in R4.")),
        ("Log Sheet (p.11)", "Blank ops log — loose-leaf, the only consumable page."),
    ]
    for comp, instr in rows:
        ws.cell(row=r, column=1, value=b).font = F_BOLD
        ws.cell(row=r, column=2, value=comp).font = F_STD
        c = ws.cell(row=r, column=3, value=instr); c.font = F_STD; c.alignment = WRAP
        for col in range(1, 4): ws.cell(row=r, column=col).border = BORDER
        r += 1
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Validation
ws = wb.create_sheet("Validation")
hdrs = ["Check", "Expected", "Actual", "Status"]
ws.append(hdrs); style_header(ws, 1, len(hdrs))
set_widths(ws, [64, 12, 12, 12])
nf = len(FAULTS) + 1
ns = len(SPEC_ROWS) + 1
checks = [
 ("Tutorial faults (R0)", 6, f'=COUNTIF(Faults!$B$2:$B${nf},"R0")'),
 ("R1 deck cards (self-contained, baseline round)", 8, f'=COUNTIF(Faults!$B$2:$B${nf},"R1")'),
 ("R2 deck cards (ring cross-refs incl. false alarm)", 12, f'=COUNTIF(Faults!$B$2:$B${nf},"R2")'),
 ("R3 deck cards (critical, two-spec)", 6, f'=COUNTIF(Faults!$B$2:$B${nf},"R3")'),
 ("R4 deck cards (novel combinations)", 4, f'=COUNTIF(Faults!$B$2:$B${nf},"R4")'),
 ("Spec values all unique (SUMPRODUCT self-count = row count)", ns - 1,
  f'=SUMPRODUCT(COUNTIF(SpecTables!$F$2:$F${ns},SpecTables!$F$2:$F${ns}))'),
 ("Buried appendix specs referenced by faults (all six)", 6,
  '=SUMPRODUCT((COUNTIF(Faults!$K$2:$K$37,{"PC-1";"WC-1";"MC-1";"TC-1";"AC-1";"CC-1"})'
  '+COUNTIF(Faults!$M$2:$M$37,{"PC-1";"WC-1";"MC-1";"TC-1";"AC-1";"CC-1"})>0)*1)'),
 ("Discrepancy cell W4-3 prints 340", 340,
  '=INDEX(SpecTables!$F:$F,MATCH("W4-3",SpecTables!$A:$A,0))'),
 ("Reserved value 290 appears in no table", 0, f'=COUNTIF(SpecTables!$F$2:$F${ns},290)'),
]
for b in SECTORS:
    checks.append((f"{b} binder demanded by other sectors' faults ≥ 2 (kit relevance)", "≥2",
                   f'=SUMIFS(SpecTables!$G$2:$G${ns},SpecTables!$B$2:$B${ns},"{b}")'))
r = 2
for name, exp, formula in checks:
    ws.cell(row=r, column=1, value=name).font = F_STD
    ws.cell(row=r, column=2, value=exp).font = F_STD
    ws.cell(row=r, column=3, value=formula).font = F_STD
    if exp == "≥2":
        st = f'=IF(C{r}>=2,"OK","CHECK")'
    else:
        st = f'=IF(C{r}={exp},"OK","CHECK")'
    ws.cell(row=r, column=4, value=st).font = F_BOLD
    for c in range(1, 5): ws.cell(row=r, column=c).border = BORDER
    r += 1
ws.cell(row=r + 1, column=1, value="All rows must read OK before printing binders or exporting faults.json.").font = F_NOTE

# Defaults to the filename export_faults.py looks for in the working directory.
OUT = Path(sys.argv[1] if len(sys.argv) > 1 else "undercity-crossref-matrix.xlsx")
wb.save(OUT)
print("saved", len(FAULTS), "faults,", len(SPEC_ROWS), "spec rows ->", OUT)
print("NEXT: recalculate the workbook, then run export_faults.py")
