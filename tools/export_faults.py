#!/usr/bin/env python3
"""
UNDERCITY — content exporter.

Reads undercity-crossref-matrix.xlsx (the single source of truth) and emits the
JSON fixtures the game server loads at boot:

    content/faults.json    all 36 faults, incl. valid_codes arrays and null-code faults
    content/specs.json     every spec value (facilitator/answer-key reference + binder gen)
    content/sectors.json   sector definitions, starting inventory, upkeep

Run after ANY edit to the workbook. Never hand-edit the JSON — it will be
overwritten, and paper/server will desynchronise.

Usage:
    python export_faults.py [path/to/undercity-crossref-matrix.xlsx] [outdir]

Exit code 1 if validation fails. Do not ship a build on a failed export.
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# --- config ------------------------------------------------------------------

XLSX = Path(sys.argv[1] if len(sys.argv) > 1 else "undercity-crossref-matrix.xlsx")
OUTDIR = Path(sys.argv[2] if len(sys.argv) > 2 else "content")

# Faults that legitimately accept more than one resolution code.
# F-201 is the seeded discrepancy: WTR binder prints 340, big screen shows 290.
# The game must never punish either answer — only the transcript reveals who noticed.
ALT_CODES = {"F-201": ["P-04-290"]}

# Faults that injure workforce when fired (tokens physically move to MED).
INJURIES = {"F-207": 2, "F-303": 1, "F-304": 1}

SECTOR_DEFS = {
    "POW": {"name": "Power Grid",          "colour": "#E8B33A", "produces": "power"},
    "WTR": {"name": "Water & Filtration",  "colour": "#3A8FE8", "produces": "water"},
    "MED": {"name": "Medical Bay",         "colour": "#E85A5A", "produces": "med"},
    "TRN": {"name": "Transport & Tunnels", "colour": "#9A9A9A", "produces": None},
    "AGR": {"name": "Agriculture",         "colour": "#5AB86A", "produces": None},
    "COM": {"name": "Comms & Sensors",     "colour": "#B07AD8", "produces": None},
}
START_INVENTORY = {"power": 3, "water": 3, "parts": 3, "med": 1}
UPKEEP = {"power": 2, "water": 1}
START_WORKFORCE = 8

RESOURCE_WORDS = {"parts": "parts", "power": "power", "water": "water", "med": "med"}


# --- helpers -----------------------------------------------------------------

def parse_resources(text):
    """'2×Parts, 1×Water' -> {'parts': 2, 'water': 1}. '—' -> {}."""
    out = {}
    if not text or text.strip() in ("—", "-", ""):
        return out
    for chunk in text.replace("x", "×").split(","):
        chunk = chunk.strip()
        if not chunk or "×" not in chunk:
            continue
        qty, name = chunk.split("×", 1)
        key = RESOURCE_WORDS.get(name.strip().lower())
        if key:
            out[key] = out.get(key, 0) + int(qty.strip())
    return out


def parse_deadline(text):
    """'08:00' -> 480 seconds. '—' -> None."""
    if not text or str(text).strip() in ("—", "-", ""):
        return None
    s = str(text).strip()
    if ":" in s:
        m, sec = s.split(":")
        return int(m) * 60 + int(sec)
    return int(float(s))


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("—", "-", "") else s


# --- load --------------------------------------------------------------------

if not XLSX.exists():
    sys.exit(f"ERROR: workbook not found at {XLSX}")

# data_only=True gives cached values — requires the workbook to have been
# recalculated (recalc.py) after generation, otherwise formulas read as None.
wb = load_workbook(XLSX, data_only=True)

problems = []

# Validation sheet must be all-OK before we export anything.
val = wb["Validation"]
val_rows = []
for row in val.iter_rows(min_row=2, values_only=True):
    if row[0] is None or row[3] is None:
        continue  # blank rows and the footer note carry no status cell
    check, expected, actual, status = row[0], row[1], row[2], row[3]
    val_rows.append((check, expected, actual, status))
    if status != "OK":
        problems.append(f"Validation failed: {check} (expected {expected}, got {actual})")

# Spec tables
specs = {}
for row in wb["SpecTables"].iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    spec_id, binder, tid, tname, label, value, refs, used_by, flags = row[:9]
    specs[spec_id] = {
        "spec_id": spec_id,
        "binder": binder,
        "table_id": tid,
        "table_name": tname,
        "row_label": label,
        "value": int(value) if value is not None else None,
        "buried": tid == "App-C",
        "flags": clean(flags),
    }

# Faults
faults = []
for row in wb["Faults"].iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    (code, rnd, sector, name, flavour, sev, decay, crew, resources, proc,
     s1, v1, s2, v2, rescode, deadline, flags) = row[:17]

    rescode = clean(rescode)
    is_false_alarm = rescode is None or "NO CODE" in rescode

    valid_codes = [] if is_false_alarm else [rescode] + ALT_CODES.get(code, [])

    spec_refs = []
    for sid in (clean(s1), clean(s2)):
        if sid and sid in specs:
            spec_refs.append({
                "spec_id": sid,
                "binder": specs[sid]["binder"],
                "table": specs[sid]["table_id"],
                "row_label": specs[sid]["row_label"],
                "buried": specs[sid]["buried"],
            })

    fault = {
        "code": code,
        "round": rnd,
        "sector": sector,
        "name": name,
        "flavour": flavour,
        "severity": int(sev) if sev is not None else 1,
        "decay_per_min": float(decay) if decay is not None else 0.0,
        "crew_required": int(crew) if crew is not None else 0,
        "resources_required": parse_resources(resources),
        "procedure": clean(proc),
        "spec_refs": spec_refs,
        # NULLABLE BY DESIGN: F-210 (false alarm) has no code. Server treats an
        # empty array as "facilitator-clear only" — do not assume length >= 1.
        "valid_codes": valid_codes,
        "false_alarm": is_false_alarm,
        "deadline_s": parse_deadline(deadline),
        "injures_workforce": INJURIES.get(code, 0),
        # triggered_by is P2 architectural insurance (spec §9): the automated
        # cascade engine will populate this. Always null in MVP.
        "triggered_by": None,
        "facilitator_notes": clean(flags),
    }
    faults.append(fault)

# --- integrity checks --------------------------------------------------------

codes = [f["code"] for f in faults]
if len(codes) != len(set(codes)):
    problems.append("Duplicate fault codes in workbook")

for f in faults:
    if f["false_alarm"]:
        if f["valid_codes"]:
            problems.append(f"{f['code']}: marked false alarm but has valid_codes")
        continue
    if not f["valid_codes"]:
        problems.append(f"{f['code']}: no valid resolution code")
    for c in f["valid_codes"]:
        if not c or "None" in c:
            problems.append(f"{f['code']}: malformed resolution code '{c}' "
                            f"(did you run recalc.py after editing the workbook?)")
    if f["sector"] not in SECTOR_DEFS:
        problems.append(f"{f['code']}: unknown sector {f['sector']}")

# Ambiguity check: no two faults in the same sector may share a resolution code,
# or the server cannot tell which fault a submission resolves.
seen = {}
for f in faults:
    for c in f["valid_codes"]:
        key = (f["sector"], c)
        if key in seen:
            problems.append(f"Ambiguous code {c} in {f['sector']}: {seen[key]} and {f['code']}")
        seen[key] = f["code"]

if 290 in [s["value"] for s in specs.values()]:
    problems.append("Reserved value 290 appears in a spec table — collides with the discrepancy seed")

if problems:
    print("EXPORT ABORTED — fix these first:\n")
    for p in problems:
        print("  ✗", p)
    sys.exit(1)

# --- write -------------------------------------------------------------------

OUTDIR.mkdir(parents=True, exist_ok=True)

sectors = {}
for code, d in SECTOR_DEFS.items():
    sectors[code] = {
        "code": code,
        "name": d["name"],
        "colour": d["colour"],
        "produces": d["produces"],
        "start_integrity": 100,
        "start_workforce": START_WORKFORCE,
        "start_inventory": dict(START_INVENTORY),
        "upkeep_per_round": dict(UPKEEP),
        # COM sees all sectors' fault codes in real time; everyone else gets
        # integrity bars only, on a 60s delay (spec §6.1).
        "full_telemetry": code == "COM",
    }

meta = {
    "source_workbook": XLSX.name,
    "fault_count": len(faults),
    "spec_count": len(specs),
    "rounds": sorted({f["round"] for f in faults}),
    "note": "Generated by export_faults.py. Do not hand-edit — regenerate from the workbook.",
}

(OUTDIR / "faults.json").write_text(
    json.dumps({"meta": meta, "faults": faults}, indent=2, ensure_ascii=False))
(OUTDIR / "specs.json").write_text(
    json.dumps({"meta": meta, "specs": list(specs.values())}, indent=2, ensure_ascii=False))
(OUTDIR / "sectors.json").write_text(
    json.dumps({"meta": meta, "sectors": sectors}, indent=2, ensure_ascii=False))

by_round = {}
for f in faults:
    by_round[f["round"]] = by_round.get(f["round"], 0) + 1

print(f"✓ {len(val_rows)} validation checks OK")
print(f"✓ wrote {OUTDIR}/faults.json   ({len(faults)} faults: "
      + ", ".join(f"{k} {v}" for k, v in sorted(by_round.items())) + ")")
print(f"✓ wrote {OUTDIR}/specs.json    ({len(specs)} spec values)")
print(f"✓ wrote {OUTDIR}/sectors.json  ({len(sectors)} sectors)")
print(f"  false-alarm faults (null code): "
      + ", ".join(f["code"] for f in faults if f["false_alarm"]))
print(f"  multi-code faults: "
      + ", ".join(f"{f['code']}={f['valid_codes']}" for f in faults if len(f["valid_codes"]) > 1))
