#!/usr/bin/env python3
r"""
UNDERCITY — binder content assembler.

Reads undercity-crossref-matrix.xlsx (source of truth) and emits
binder_content.json, which build_binders.js renders into six DOCX binders.

Pipeline: matrix -> recalc -> export_faults.py (server JSON)
                           \-> assemble_binders.py -> build_binders.js (paper)

Both branches read the SAME cells, so paper and server cannot disagree.

CRITICAL CONTENT RULES ENFORCED HERE:
  1. A binder NEVER prints a full resolution code. It prints the procedure ref,
     the resource/crew cost, and WHERE to fetch each spec value. The team
     assembles the code by talking to other sectors — that assembly IS the game.
  2. A binder NEVER prints another sector's spec values. Only its own tables.
  3. WTR Table W-4 row 3 prints 340. The big screen shows 290. Do not reconcile.
  4. Appendix C gets NO index entry. It is findable only by reading the binder.
  5. AGR's F-210 index entry points to verification, not a procedure (false alarm).
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(sys.argv[1] if len(sys.argv) > 1 else "undercity-crossref-matrix.xlsx")
OUT = Path(sys.argv[2] if len(sys.argv) > 2 else "binder_content.json")

SECTOR_INFO = {
    "POW": {
        "name": "Power Grid",
        "colour": "E8B33A",
        "motto": "The city runs on what we hold.",
        "mission": (
            "Power Grid operates HAVEN-9's five geothermal turbines, the ring main, and three "
            "sub-grids. Every other sector draws from you. Turbine output is finite and the Core "
            "coupling is the single point of failure for the entire city — if the ring main drops, "
            "life support in the Medical Bay has roughly four minutes of reserve. Your standing "
            "orders are to maintain ring-main stability, honour agreed power transfers, and keep "
            "Turbine C's bearing temperature inside tolerance. You are permitted to shed load to "
            "protect the ring main. You are not permitted to shed load to the Medical Bay without "
            "Council authorisation. Power cannot be manufactured faster than the turbines allow, "
            "so every commitment you make to another sector is a commitment taken from somewhere else."
        ),
        "produces": "Power Cells (⚡) — 3 per cycle, distributed at your discretion",
    },
    "WTR": {
        "name": "Water & Filtration",
        "colour": "3A8FE8",
        "motto": "Nothing here is wasted twice.",
        "mission": (
            "Water & Filtration draws from the deep aquifer, runs four pump stations, and operates "
            "the grey-water reclaim plant that makes a closed city possible. You also supply turbine "
            "coolant to Power Grid — which makes you and POW mutual hostages, a fact both sectors "
            "are expected to manage like adults. The Lower Reservoir wall has been on the watch list "
            "for two years. Your standing orders are to maintain reservoir pressure inside rated "
            "limits, keep reclaim output within Medical Bay's bio-thresholds, and never run a pump "
            "station dry. Filtration failures are not immediately visible to anyone else in the city, "
            "which means the decision to report a problem early is yours alone to make."
        ),
        "produces": "Water Units (💧) — 3 per cycle, distributed at your discretion",
    },
    "MED": {
        "name": "Medical Bay",
        "colour": "E85A5A",
        "motto": "We count in people, not units.",
        "mission": (
            "Medical Bay runs three wards, an isolation wing, a triage bay, and the surgical suite, "
            "and holds the city's only cold-chain store. You treat every injured worker HAVEN-9 has: "
            "when another sector takes casualties, their people arrive at your door and cannot return "
            "to work until you release them. You also set the bio-thresholds that constrain Water's "
            "reclaim plant. Your standing orders are to maintain sterile atmosphere in all wards, "
            "protect the cold chain, and keep life support powered without exception. You have the "
            "strongest moral claim in any triage argument and the weakest ability to generate the "
            "resources you need. Both of those are true at once."
        ),
        "produces": "Med Supplies (⚕) — 1 per cycle; restores injured Workforce",
    },
    "TRN": {
        "name": "Transport & Tunnels",
        "colour": "9A9A9A",
        "motto": "Everything moves through us.",
        "mission": (
            "Transport & Tunnels maintains four galleries, the freight spur, the rail network, and "
            "every blast door in HAVEN-9. You produce nothing, and nothing reaches anyone without "
            "you: every resource transfer in this city is stamped by Transport before it takes "
            "effect. That makes you the quietest form of power in HAVEN-9 and the easiest to resent. "
            "Your standing orders are to keep Tunnels A through D load-rated and passable, maintain "
            "rail voltage inside spec, and stamp agreed transfers promptly. Tunnel B has had roof "
            "movement logged in three consecutive inspections. A collapse there injures crew — yours."
        ),
        "produces": "Nothing. You control transfer capacity — stamp required on every chit.",
    },
    "AGR": {
        "name": "Agriculture",
        "colour": "5AB86A",
        "motto": "Slow problems kill slowly.",
        "mission": (
            "Agriculture runs four hydroponic bays, the seedling bay, and the seed vault that "
            "represents HAVEN-9's only path to a future beyond the current generation. You feed the "
            "city. Your failures are the least urgent-looking and the most expensive: a grow-lamp "
            "array lost today shows up as a workforce efficiency problem three cycles from now, long "
            "after the argument about who gets power has been settled without you. Your standing "
            "orders are to maintain nutrient mix ratios, keep all arrays on cycle, and protect the "
            "vault environment. You will spend this shift arguing for resources against sectors whose "
            "emergencies are louder than yours. Prepare that argument before you need it."
        ),
        "produces": "Food — sustains Workforce efficiency city-wide",
    },
    "COM": {
        "name": "Comms & Sensors",
        "colour": "B07AD8",
        "motto": "We see it first. What we do next is the question.",
        "mission": (
            "Comms & Sensors operates the sensor grid, the relay network, and the uplink. Your "
            "telemetry covers every sector in HAVEN-9, which means you routinely see a fault before "
            "the sector suffering it does. You hold the City Charter. Nothing in your standing orders "
            "tells you how quickly to pass on what you see, or to whom, or what to do when your "
            "readings contradict another sector's instruments — those are judgement calls, and they "
            "are yours. Your standing orders are to maintain sensor calibration, keep the relay "
            "network powered, and preserve the uplink. When the grid goes down, the city is blind, "
            "and every other sector is arguing from memory."
        ),
        "produces": "Telemetry — you alone see other sectors' live fault data",
    },
}

START_INVENTORY = "3 ⚡ Power Cells · 3 💧 Water Units · 3 🔧 Spare Parts · 1 ⚕ Med Supply"
UPKEEP = "2 ⚡ Power Cells + 1 💧 Water Unit per cycle"

wb = load_workbook(XLSX, data_only=True)

# ---- specs ------------------------------------------------------------------
specs = {}
for row in wb["SpecTables"].iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    specs[row[0]] = {"id": row[0], "binder": row[1], "table_id": row[2],
                     "table_name": row[3], "row_label": row[4], "value": row[5],
                     "buried": row[2] == "App-C"}
if specs["W4-3"]["value"] != 340:
    sys.exit("ABORT: discrepancy seed W4-3 is not 340 — check the workbook.")

# ---- faults -----------------------------------------------------------------
faults = []
for row in wb["Faults"].iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    faults.append({
        "code": row[0], "round": row[1], "sector": row[2], "name": row[3],
        "flavour": row[4], "severity": row[5], "crew": row[7],
        "resources": row[8], "procedure": row[10 - 1],
        "spec1": row[10], "spec2": row[12],
        "deadline": row[15] if row[15] not in ("—", None) else None,
        "notes": row[16] or "",
    })

# ---- escalate entries (parsed from CrossrefMap) -----------------------------
escalate = {}
for row in wb["CrossrefMap"].iter_rows(min_row=1, values_only=True):
    if row[0] and isinstance(row[0], str) and row[0].endswith("binder index also lists:"):
        binder = row[0].split()[0]
        entries = []
        for part in (row[1] or "").split("·"):
            m = re.search(r"(F-\d+)\s*→\s*escalate to\s*([A-Z]{3})", part)
            if m:
                entries.append({"code": m.group(1), "owner": m.group(2)})
        escalate[binder] = entries

# ---- assemble ---------------------------------------------------------------
binders = {}
for code, info in SECTOR_INFO.items():
    own = [f for f in faults if f["sector"] == code]

    index_rows = []
    for f in sorted(own, key=lambda x: x["code"]):
        if f["procedure"] in (None, "—"):
            index_rows.append({
                "code": f["code"], "name": f["name"],
                "action": "VERIFY TELEMETRY WITH COM BEFORE COMMITTING RESOURCES",
                "own": True, "no_procedure": True,
            })
        else:
            index_rows.append({
                "code": f["code"], "name": f["name"],
                "action": f"Procedure {f['procedure']}", "own": True, "no_procedure": False,
            })
    for e in escalate.get(code, []):
        index_rows.append({
            "code": e["code"], "name": "Not a " + code + " system fault",
            "action": f"ESCALATE TO {e['owner']}", "own": False, "no_procedure": False,
        })

    procedures = []
    for f in sorted([x for x in own if x["procedure"] not in (None, "—")],
                    key=lambda x: x["procedure"]):
        steps, sources, nparts = [], [], 0
        for sid in (f["spec1"], f["spec2"]):
            if sid and sid in specs:
                nparts += 1
                s = specs[sid]
                where = ("YOUR Appendix C" if s["buried"] and s["binder"] == code
                         else f"{s['binder']} Manual, Table {s['table_id']}"
                         if s["binder"] != code else f"YOUR Table {s['table_id']}")
                sources.append({"where": where, "row_label": s["row_label"],
                                "foreign": s["binder"] != code, "buried": s["buried"]})
        fmt = f["procedure"] + "-[VALUE]" + ("-[VALUE 2]" if nparts == 2 else "")
        steps.append(f"Confirm the fault code on the alert card matches {f['code']}.")
        steps.append(f"Assign crew: {f['crew']} worker(s) minimum. Fewer will not hold the isolation.")
        steps.append(f"Stage materials: {f['resources']}.")
        for i, s in enumerate(sources, 1):
            verb = ("Obtain" if s["foreign"] else "Read off")
            steps.append(f"{verb} the {s['row_label']} value from {s['where']}." +
                         (" This value is not held in this binder." if s["foreign"] else ""))
        steps.append(f"Enter the resolution code on the sector console in the format {fmt}, "
                     "substituting the value(s) above. Values are three digits.")
        steps.append("If the console rejects the entry, re-verify the source table before "
                     "resubmitting. Three consecutive rejections lock the console for 20 seconds.")
        procedures.append({
            "id": f["procedure"], "fault_code": f["code"], "title": f["name"],
            "resources": f["resources"], "crew": f["crew"],
            "deadline": f["deadline"], "format": fmt,
            "sources": sources, "steps": steps,
            "severity": f["severity"],
        })

    tables = {}
    for s in specs.values():
        if s["binder"] != code or s["buried"]:
            continue
        tables.setdefault(s["table_id"], {"id": s["table_id"], "name": s["table_name"], "rows": []})
        tables[s["table_id"]]["rows"].append({"label": s["row_label"], "value": s["value"]})

    appendix = next(s for s in specs.values() if s["binder"] == code and s["buried"])

    binders[code] = {
        "code": code, "name": info["name"], "colour": info["colour"],
        "motto": info["motto"], "mission": info["mission"], "produces": info["produces"],
        "upkeep": UPKEEP, "start_inventory": START_INVENTORY,
        "index_rows": index_rows,
        "procedures": procedures,
        "tables": sorted(tables.values(), key=lambda t: t["id"]),
        "appendix": {"row_label": appendix["row_label"], "value": appendix["value"]},
    }

# ---- leak check: no binder may print another binder's values -----------------
for code, b in binders.items():
    printed = {r["value"] for t in b["tables"] for r in t["rows"]} | {b["appendix"]["value"]}
    for sid, s in specs.items():
        if s["binder"] != code and s["value"] in printed:
            sys.exit(f"ABORT: {code} binder prints {s['value']}, which belongs to {s['binder']} ({sid})")

OUT.write_text(json.dumps({"binders": binders}, indent=2, ensure_ascii=False))
print(f"✓ {OUT}")
for c, b in binders.items():
    print(f"  {c}: {len(b['index_rows'])} index rows "
          f"({sum(1 for r in b['index_rows'] if not r['own'])} escalate), "
          f"{len(b['procedures'])} procedures, {len(b['tables'])} tables, "
          f"appendix '{b['appendix']['row_label']}' = {b['appendix']['value']}")
print("✓ leak check passed — no binder prints another sector's spec values")
